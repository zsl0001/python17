import sys

sys.path.append("..")

from models import db, TMSOrderIndex, TMSDevice
from my_db.Models import Position, Log, Status
import time
import datetime
from mongoengine.queryset.visitor import Q
import openpyxl
from geopy.distance import geodesic

s_time = '2020-12-01 00:00:00'
e_time = '2020-12-31 23:59:59'


def get_location_time(t):
    if len(str(t)) > 19:
        t = str(t).split('.')[0]
    t = datetime.datetime.strptime(str(t), '%Y-%m-%d %H:%M:%S') + datetime.timedelta(hours=8)
    l_t = t.strftime("%Y-%m-%d %H:%M:%S")
    return t


def local2utc(local_st):
    """本地时间转UTC时间（-8: 00）"""
    if len(str(local_st)) > 19:
        local_st = str(local_st).split('.')[0]
    local_st = datetime.datetime.strptime(str(local_st), '%Y-%m-%d %H:%M:%S')
    time_struct = time.mktime(local_st.timetuple())
    utc_st = datetime.datetime.utcfromtimestamp(time_struct)
    return utc_st


def get_Utc_time(t):
    if len(str(t)) > 19:
        t = str(t).split('.')[0]
    a = datetime.datetime.strptime(str(t), "%Y-%m-%d %H:%M:%S")
    o = datetime.timedelta(hours=8)
    return a - o


def get_day3_time(t):
    if len(str(t)) > 19:
        t = str(t).split('.')[0]
    a = datetime.datetime.strptime(str(t), "%Y-%m-%d %H:%M:%S")
    a = a + datetime.timedelta(days=3)
    o = datetime.timedelta(hours=8)
    return a - o


def get_day4_time(t):
    if len(str(t)) > 19:
        t = str(t).split('.')[0]
    a = datetime.datetime.strptime(str(t), "%Y-%m-%d %H:%M:%S")
    a = a + datetime.timedelta(days=3)
    return a


def set_time(imei, start_time, sign_time=None):  # 将时间按3小时一次切割查询
    l = []
    dat = {'start_time': start_time,
           'end_time': sign_time,
           'imei': imei
           }
    # print('start_time{},sign_time{}'.format(start_time, sign_time))
    t = str(start_time)
    while t <= str(sign_time):
        dat['start_time'] = t
        dat['end_time'] = str(datetime.datetime.strptime(t, '%Y-%m-%d %H:%M:%S') + datetime.timedelta(hours=3))
        t = str(datetime.datetime.strptime(t, '%Y-%m-%d %H:%M:%S') + datetime.timedelta(hours=3))
        l.append(dat.copy())
    l[-1]['end_time'] = sign_time
    return l


def get_position(data):
    my_list = []
    st = local2utc(str(data['start_time']))
    ed = local2utc(str(data['end_time']))
    res = Position.objects.filter((Q(time__gte=st) & Q(time__lte=ed) & Q(devId=str(data['imei'])))).order_by('time')
    res2 = Position.objects.filter(Q(time__lte=ed) & Q(devId=str(data['imei']))).order_by('-time').limit(1)
    for i in res:
        p_data = {
            'devId': i.devId,
            "longitude": i.longitude,
            "latitude": i.latitude,
            "time": get_location_time(i.time)
        }
        my_list.append(p_data)
    for i in res2:
        p_data2 = {
            'devId': i.devId,
            "longitude": i.longitude,
            "latitude": i.latitude,
            "time": get_location_time(i.time)
        }
    return my_list, p_data2


# def get_all_location(imei, start_time, end_time=None):
#     # sign_time = get_day4_time(end_time)
#     l = set_time(imei=imei, start_time=start_time, sign_time=end_time)
#     m = []
#     for i in l:
#         m = get_position(i) + m
#     return m


wb = openpyxl.load_workbook('test3.xlsx')
sheet = wb['Sheet1']
max_row = sheet.max_row
for row in range(2, max_row + 1):
    # 出发时间
    s_time = sheet.cell(row, 2).value
    # 到达时间
    e_time = sheet.cell(row, 3).value
    # 设备码
    imei = sheet.cell(row, 8).value
    data = {'start_time':s_time,'end_time':e_time,'imei':imei}
    a, b = get_position(data)
    sheet.cell(row, 17).value = len(a)
    if len(a):
        sheet.cell(row, 18).value = b['time']

wb.save('test4.xlsx')
