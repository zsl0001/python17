import sys

sys.path.append("..")

from models import db, TMSOrderIndex, TMSDevice
from my_db.Models import Position, Log, Status
import time
import datetime
from mongoengine.queryset.visitor import Q
import openpyxl
from geopy.distance import geodesic

s_time = '2020-11-01 00:00:00'
e_time = '2020-11-30 00:00:00'


def get_location_time(t):
    if len(str(t)) > 19:
        t = str(t).split('.')[0]
    t = datetime.datetime.strptime(str(t), '%Y-%m-%d %H:%M:%S') + datetime.timedelta(hours=8)
    l_t = t.strftime("%Y-%m-%d %H:%M:%S")
    return t


def local2utc(local_st):
    """本地时间转UTC时间（-8: 00）"""
    if len(str(local_st)) > 19:
        local_st = str(local_st).split('.')[0]
    local_st = datetime.datetime.strptime(str(local_st), '%Y-%m-%d %H:%M:%S')
    time_struct = time.mktime(local_st.timetuple())
    utc_st = datetime.datetime.utcfromtimestamp(time_struct)
    return utc_st


def get_Utc_time(t):
    if len(str(t)) > 19:
        t = str(t).split('.')[0]
    a = datetime.datetime.strptime(str(t), "%Y-%m-%d %H:%M:%S")
    o = datetime.timedelta(hours=8)
    return a - o


def get_day3_time(t):
    if len(str(t)) > 19:
        t = str(t).split('.')[0]
    a = datetime.datetime.strptime(str(t), "%Y-%m-%d %H:%M:%S")
    a = a + datetime.timedelta(days=3)
    o = datetime.timedelta(hours=8)
    return a - o


def get_day4_time(t):
    if len(str(t)) > 19:
        t = str(t).split('.')[0]
    a = datetime.datetime.strptime(str(t), "%Y-%m-%d %H:%M:%S")
    a = a + datetime.timedelta(days=3)
    return a


def set_time(imei, start_time, sign_time=None):  # 将时间按3小时一次切割查询
    l = []
    dat = {'start_time': start_time,
           'end_time': sign_time,
           'imei': imei
           }
    print('start_time{},sign_time{}'.format(start_time, sign_time))
    t = str(start_time)
    while t <= str(sign_time):
        dat['start_time'] = t
        dat['end_time'] = str(datetime.datetime.strptime(t, '%Y-%m-%d %H:%M:%S') + datetime.timedelta(hours=3))
        t = str(datetime.datetime.strptime(t, '%Y-%m-%d %H:%M:%S') + datetime.timedelta(hours=3))
        l.append(dat.copy())
    l[-1]['end_time'] = sign_time
    return l


def get_position(data):
    my_list = []
    st = local2utc(str(data['start_time']))
    ed = local2utc(str(data['end_time']))
    res = Position.objects.filter((Q(time__gte=st) & Q(time__lte=ed) & Q(devId=str(data['imei'])))).order_by('time')
    for i in res:
        p_data = {
            'devId': i.devId,
            "longitude": i.longitude,
            "latitude": i.latitude,
            "time": get_location_time(i.time)
        }
        my_list.append(p_data)
    return my_list


def get_all_location(imei, start_time, end_time=None):
    sign_time = get_day4_time(end_time)
    l = set_time(imei=imei, start_time=start_time, sign_time=sign_time)
    m = []
    for i in l:
        m = get_position(i) + m
    return m


wb = openpyxl.load_workbook('all2.xlsx')
sheet = wb['Sheet1']
max_row = sheet.max_column
print(sheet.cell(1, 1).value)
# for row in range(2, max_row + 1):
#     imei = sheet.cell(row, 2).value
#     stime = sheet.cell(row, 3).value
my_res = db.session.query(TMSOrderIndex.Index_PactCode, TMSOrderIndex.Index_FromTime,
                          TMSOrderIndex.Index_ToTime, TMSOrderIndex.Index_SignTime,
                          TMSOrderIndex.Index_DeviceBindingTime, TMSOrderIndex.Index_FromLocation,
                          TMSOrderIndex.Index_ToLocation
                          ).filter(TMSOrderIndex.Index_SrcClass == 1,
                                   # TMSOrderIndex.Index_PactCode == '4500330063',
                                   TMSOrderIndex.Index_Status != 32,
                                   TMSOrderIndex.Index_FromTime >= s_time,
                                   TMSOrderIndex.Index_ToTime <= e_time,
                                   TMSDevice.Device_IMEICode == TMSOrderIndex.Index_DeviceCode,
                                   TMSDevice.Device_Type.between(2, 3),
                                   TMSOrderIndex.Index_CreatorCompanyID.in_([1925])).order_by(
    TMSOrderIndex.Index_ToTime)
row = 1
for i in my_res.all():
    row = row + 1
    my_list = []
    # 合同号
    print('正在处理合同号为{}的订单'.format(i[0]))
    sheet.cell(row, 1).value = i[0]
    # 预计发货时间
    sheet.cell(row, 2).value = i[1]
    # 预计到货时间
    sheet.cell(row, 3).value = i[2]
    # 签收时间
    sheet.cell(row, 4).value = i[3]
    # 签收时距离发货时间时长
    s_f = round((i[3] - i[1]).total_seconds() / 3600, 2)
    sheet.cell(row, 5).value = s_f
    # 设备绑定时间
    sheet.cell(row, 7).value = i[4]
    # 签收时间距离绑定时间时长
    s_b = round((i[3] - i[4]).total_seconds() / 3600, 2)
    sheet.cell(row, 8).value = s_b
    # 出发地和目的地总距离
    if len(str(i[-1])) >= 10:
        end_loca = (eval(str(i[-1]).split(',')[1]), eval(str(i[-1]).split(',')[0]))
    if len(str(i[-2])) >= 10:
        start_loca = (eval(str(i[-2]).split(',')[1]), eval(str(i[-2]).split(',')[0]))
    all_dis = round(geodesic(end_loca, start_loca).km, 2)
    sheet.cell(row, 10).value = all_dis
    # 时速1
    sheet.cell(row, 6).value = all_dis/s_f
    # 时速2
    sheet.cell(row, 9).value = all_dis/s_b

wb.save('all3.xlsx')
