import sys

sys.path.append("..")

from models import db, TMSOrderIndex, TMSDevice
import openpyxl

wb = openpyxl.load_workbook('SIM.xlsx')
sheet = wb['Sheet1']
max_row = sheet.max_row
print(sheet.cell(1, 1).value)
for i in range(2, max_row):
    imei = sheet.cell(i, 3).value
    phone = sheet.cell(i, 6).value
    res = db.session.query(TMSDevice).filter(TMSDevice.Device_IMEICode == imei).all()
    for k in res:
        k.Device_PhoneNo = phone
        print(k.Device_PhoneNo)
        try:
            db.session.commit()
        except Exception as e:
            print(e)
