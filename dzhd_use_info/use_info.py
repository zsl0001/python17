import sys

sys.path.append("..")

from models import db, TMSOrderIndex, TMSDevice
from my_db.Models import Position, Log, Status
import time
import datetime
from mongoengine.queryset.visitor import Q
import openpyxl
from geopy.distance import geodesic

s_time = '2020-12-29 00:00:00'
e_time = '2021-01-04 00:00:00'


def get_location_time(t):
    if len(str(t)) > 19:
        t = str(t).split('.')[0]
    t = datetime.datetime.strptime(str(t), '%Y-%m-%d %H:%M:%S') + datetime.timedelta(hours=8)
    l_t = t.strftime("%Y-%m-%d %H:%M:%S")
    return t


def local2utc(local_st):
    """本地时间转UTC时间（-8: 00）"""
    if len(str(local_st)) > 19:
        local_st = str(local_st).split('.')[0]
    local_st = datetime.datetime.strptime(str(local_st), '%Y-%m-%d %H:%M:%S')
    time_struct = time.mktime(local_st.timetuple())
    utc_st = datetime.datetime.utcfromtimestamp(time_struct)
    return utc_st


def get_Utc_time(t):
    if len(str(t)) > 19:
        t = str(t).split('.')[0]
    a = datetime.datetime.strptime(str(t), "%Y-%m-%d %H:%M:%S")
    o = datetime.timedelta(hours=8)
    return a - o


def get_day3_time(t):
    if len(str(t)) > 19:
        t = str(t).split('.')[0]
    a = datetime.datetime.strptime(str(t), "%Y-%m-%d %H:%M:%S")
    a = a + datetime.timedelta(days=3)
    o = datetime.timedelta(hours=8)
    return a - o


def get_day4_time(t):
    if len(str(t)) > 19:
        t = str(t).split('.')[0]
    a = datetime.datetime.strptime(str(t), "%Y-%m-%d %H:%M:%S")
    a = a + datetime.timedelta(days=3)
    return a


def set_time(imei, start_time, sign_time=None):  # 将时间按3小时一次切割查询
    l = []
    dat = {'start_time': start_time,
           'end_time': sign_time,
           'imei': imei
           }
    print('start_time{},sign_time{}'.format(start_time, sign_time))
    t = str(start_time)
    while t <= str(sign_time):
        dat['start_time'] = t
        dat['end_time'] = str(datetime.datetime.strptime(t, '%Y-%m-%d %H:%M:%S') + datetime.timedelta(hours=3))
        t = str(datetime.datetime.strptime(t, '%Y-%m-%d %H:%M:%S') + datetime.timedelta(hours=3))
        l.append(dat.copy())
    l[-1]['end_time'] = sign_time
    return l


def get_position(data):
    my_list = []
    st = local2utc(str(data['start_time']))
    ed = local2utc(str(data['end_time']))
    res = Position.objects.filter((Q(time__gte=st) & Q(time__lte=ed) & Q(devId=str(data['imei'])))).order_by('time')
    for i in res:
        p_data = {
            'devId': i.devId,
            "longitude": i.longitude,
            "latitude": i.latitude,
            "time": get_location_time(i.time)
        }
        my_list.append(p_data)
    return my_list


def get_all_location(imei, start_time, end_time=None):
    sign_time = get_day4_time(end_time)
    l = set_time(imei=imei, start_time=start_time, sign_time=sign_time)
    m = []
    for i in l:
        m = get_position(i) + m
    return m


wb = openpyxl.load_workbook('test1.xlsx')
sheet = wb['Sheet1']
max_row = sheet.max_column
print(sheet.cell(1, 1).value)
# for row in range(2, max_row + 1):
#     imei = sheet.cell(row, 2).value
#     stime = sheet.cell(row, 3).value
my_res = db.session.query(TMSOrderIndex.Index_PactCode, TMSOrderIndex.Index_FromTime,
                          TMSOrderIndex.Index_ToTime, TMSOrderIndex.Index_SignTime,
                          TMSOrderIndex.Index_DeviceCode, TMSOrderIndex.Index_DeviceBindingTime,
                          TMSOrderIndex.Index_Status, TMSDevice.Device_Type, TMSOrderIndex.Index_FromLocation,
                          TMSOrderIndex.Index_CreatorCompanyName, TMSDevice.Device_CompanyID,
                          TMSOrderIndex.Index_ToLocation
                          ).filter(TMSOrderIndex.Index_SrcClass == 1,
                                   # TMSOrderIndex.Index_PactCode == '4500343553',
                                   TMSOrderIndex.Index_Status != 32,
                                   TMSOrderIndex.Index_FromTime >= s_time,
                                   TMSOrderIndex.Index_FromTime <= e_time,
                                   TMSDevice.Device_IMEICode == TMSOrderIndex.Index_DeviceCode,
                                   TMSDevice.Device_Type.between(2, 3),
                                   TMSDevice.Device_CompanyID.in_([1928]),
                                   TMSOrderIndex.Index_CreatorCompanyID.in_([1925])).order_by(
    TMSOrderIndex.Index_FromTime)
row = 1
for i in my_res.all():
    row = row + 1
    my_list = []
    print('正在处理合同号为{}的订单'.format(i[0]))

    sheet.cell(row, 1).value = i[0]
    # 预计发货时间
    sheet.cell(row, 2).value = i[1]
    # 预计到货时间
    sheet.cell(row, 3).value = i[2]
    # 签收时间
    sheet.cell(row, 4).value = i[3]
    # 设备绑定时间 conda install kivy -c conda-forge

    sheet.cell(row, 5).value = i[5]
    # 创建公司
    sheet.cell(row, 15).value = i[8]
    # 线上承运方名字
    sheet.cell(row, 16).value = i[10]
    # 绑定时电量 351608086036972
    if i[7] == 3:  # 3代设备 取 status Q(time__gte=st) &
        status_res = Status.objects.filter(
            (Q(date__gte=get_Utc_time(str(i[5]))) & Q(date__lte=get_Utc_time(str(i[2]))) & Q(devId=i[4]))).order_by(
            '-date').first()
        if status_res:
            sheet.cell(row, 6).value = status_res['lithium']
            sheet.cell(row, 7).value = get_location_time(str(status_res['date']))
        else:
            sheet.cell(row, 6).value = '无历史数据'
            sheet.cell(row, 7).value = '无历史数据'
    else:  # 取 log 表
        # 绑定时电量上传时间
        log_res = Log.objects.filter(
            (Q(time__gte=get_Utc_time(str(i[5]))) & Q(time__lte=get_Utc_time(str(i[2]))) & Q(imei=i[4]))).order_by(
            'time').first()
        if log_res:
            print(log_res['content'].split(',')[-1])
            sheet.cell(row, 6).value = log_res['content'].split(',')[-1]
            sheet.cell(row, 7).value = get_location_time(str(log_res['time']))
        else:
            sheet.cell(row, 6).value = '无历史数据'
            sheet.cell(row, 7).value = '无历史数据'
    sheet.cell(row, 8).value = i[4]
    # 签收时电量  sheet.cell(row, 10).value
    if i[6] == 2:  # 待签收
        sheet.cell(row, 9).value = '订单未签收'
        sheet.cell(row, 10).value = '订单未签收'
    else:
        sheet.cell(row, 9).value = '订单已签收'
        if i[7] == 3:  # 3代设备 取 status
            status_res = Status.objects.filter(
                (Q(date__gte=get_Utc_time(str(i[1]))) & Q(date__lte=get_Utc_time(str(i[3]))) & Q(devId=i[4]))).order_by(
                '-date').first()
            if status_res:
                sheet.cell(row, 10).value = status_res['lithium']
            else:
                sheet.cell(row, 10).value = '出发时间到签收时间内无数据'
        else:  # 取 log 表
            # 绑定时电量上传时间
            log_res = Log.objects.filter(
                (Q(time__gte=get_Utc_time(str(i[1]))) & Q(time__lte=get_Utc_time(str(i[3]))) & Q(imei=i[4]))).order_by(
                'time').first()
            if log_res:
                sheet.cell(row, 10).value = log_res['content'].split(',')[-1]
            else:
                sheet.cell(row, 10).value = '出发时间到签收时间内无数据'
    # 是否中途断点 ((Q(date__gte=start) & Q(date__lte=end) & Q(devId=imei))) i[1]发货时间 i[2]到货时间 i[3]签收时间
    if i[6] == 2:  # 待签收
        p1_res = Position.objects.filter(
            (Q(time__gte=get_Utc_time(str(i[1]))) & Q(time__lte=get_Utc_time(str(i[2]))) & Q(devId=i[4]))).count()
        if p1_res:
            p2_res = Position.objects.filter(
                (Q(time__gte=get_day3_time(str(i[1]))) & Q(time__lte=get_Utc_time(str(i[2]))) & Q(
                    devId=i[4]))).count()
            if p2_res:
                all_data = get_all_location(imei=i[4], start_time=i[1], end_time=i[2])
                if len(str(i[-1])) >= 10:
                    end_loca = (eval(str(i[-1]).split(',')[1]), eval(str(i[-1]).split(',')[0]))
                if len(str(i[8])) >= 10:
                    start_loca = (eval(str(i[8]).split(',')[1]), eval(str(i[8]).split(',')[0]))
                    all_dis = round(geodesic(end_loca, start_loca).km, 2)
                    for dat_1 in all_data:
                        s_laca = (dat_1['latitude'], dat_1['longitude'])
                        my_list.append(round(geodesic(s_laca, end_loca).km, 2))
                    if my_list:
                        min_dis = min(my_list)
                        sheet.cell(row, 11).value = '截止到货期后3天内，距离目的地最短距离为{}'.format(min_dis)
                        if min_dis / all_dis >= 0.5:
                            sheet.cell(row, 17).value = '不足一半'
                        if min_dis / all_dis <= 0.5:
                            sheet.cell(row, 17).value = '超过一半'
                    else:
                        sheet.cell(row, 11).value = '截止到货期后3天内无定位数据！'
                else:
                    sheet.cell(row, 11).value = '经纬度缺失'

            else:
                sheet.cell(row, 12).value = '截止到货期后3天内无定位数据！'
        else:
            sheet.cell(row, 11).value = '截止到货期后3天内无定位数据！'
            sheet.cell(row, 12).value = '截止到货期后3天内无定位数据！'
    else:
        p1_res = Position.objects.filter(
            (Q(time__gte=get_Utc_time(str(i[1]))) & Q(time__lte=get_Utc_time(str(i[3]))) & Q(devId=i[4]))).count()
        if p1_res:
            all_data = get_all_location(imei=i[4], start_time=i[1], end_time=i[3])
            if i[-1]:
                # print(i[0])
                end_loca = (eval(str(i[-1]).split(',')[1]), eval(str(i[-1]).split(',')[0]))
            if len(str(i[8])) >= 10:
                start_loca = (eval(str(i[8]).split(',')[1]), eval(str(i[8]).split(',')[0]))
                all_dis = round(geodesic(end_loca, start_loca).km, 2)
                for dat_1 in all_data:
                    s_laca = (dat_1['latitude'], dat_1['longitude'])
                    my_list.append(round(geodesic(s_laca, end_loca).km, 2))
                if my_list:
                    min_dis = min(my_list)
                    sheet.cell(row, 11).value = '截止签收时间，距离目的地最短距离为{}'.format(min(my_list))
                    if min_dis / all_dis >= 0.5:
                        sheet.cell(row, 17).value = '不足一半'
                    if min_dis / all_dis <= 0.5:
                        sheet.cell(row, 17).value = '超过一半'
                else:
                    sheet.cell(row, 11).value = '截止签收时间内无定位数据！'
            else:
                sheet.cell(row, 11).value = '经纬度缺失'
        else:
            sheet.cell(row, 11).value = '截止签收时间内无定位数据！'
            sheet.cell(row, 12).value = '截止签收时间内无定位数据！'
    # 是否全程无点截止到货期后3天内无定位数据！
    # 订单签收状
    sheet.cell(row, 13).value = ('待签收' if i[6] == 2 else '已签收')
    # 预计到货时间电量
    if i[7] == 3:  # 3代设备 取 status
        status_res = Status.objects.filter((Q(date__lte=get_Utc_time(str(i[2]))) & Q(devId=i[4]))).order_by(
            '-date').first()
        if status_res:
            sheet.cell(row, 14).value = status_res['lithium']
        else:
            sheet.cell(row, 14).value = '无历史数据'
    else:  # 取 log 表
        # 绑定时电量上传时间
        log_res = Log.objects.filter((Q(time__lte=get_Utc_time(str(i[2]))) & Q(imei=i[4]))).order_by(
            '-date').first()
        if log_res:
            sheet.cell(row, 14).value = log_res['content'].split(',')[-1]
        else:
            sheet.cell(row, 14).value = '无历史数据'

wb.save('test2.xlsx')