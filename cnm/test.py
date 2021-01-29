import sys

sys.path.append("..")
import datetime
import json
import openpyxl

import pymongo
import requests


def db_electric(imei, startTime):
    sql = {'imei': imei, 'time': {'$gte': startTime}}
    print(sql)
    result = collection.find(sql).sort("time", 1).limit(1)
    data = {
        'imei': '',
        'content': '',
        'time': ''
    }
    for i in result:
        data['imei'] = i['imei']
        data['content'] = i['content'].split(',')[-1]
        print(i['content'].split(',')[-1])
        da = i['time'] + datetime.timedelta(hours=8)
        data['time'] = da.strftime("%Y-%m-%d %H:%M:%S")
    return data


wb = openpyxl.load_workbook('1111.xlsx')
sheet = wb['Sheet2']
max_row = sheet.max_row
for row in range(2, max_row + 1):
    client = pymongo.MongoClient(
        'mongodb://er_user:vE0UmSo1fV3q@dds-uf6a4f0597f9e3041.mongodb.rds.aliyuncs.com:3717,dds-uf6a4f0597f9e3042.mongodb.rds.aliyuncs.com:3717/er?replicaSet=mgset-13011481')
    db = client.er
    collection = db.log
    imei = sheet.cell(row, 2).value
    stime = sheet.cell(row, 3).value
    stime = datetime.datetime.strptime(str(stime), "%Y-%m-%d %H:%M:%S.%f")
    t = stime + +datetime.timedelta(hours=-8)
    sql = {'imei': str(imei), 'time': {'$gte': t}}
    result = collection.find(sql).sort("time", 1).limit(1)
    for i in result:
        sheet.cell(row, 4).value = i['content'].split(',')[-1]
        sheet.cell(row, 5).value = i['time'] + datetime.timedelta(hours=8)
wb.save('test.xlsx')
