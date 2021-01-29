import sys

sys.path.append("..")

import openpyxl

wb1 = openpyxl.load_workbook('test2.xlsx')
sheet1 = wb1['Sheet1']
max_row1 = sheet1.max_row
print(max_row1)
wb2 = openpyxl.load_workbook('12-131.xlsx')
sheet2 = wb2['Sheet2']
max_row2 = sheet2.max_row
print(max_row2)
for row1 in range(2, max_row1 + 1):
    for row2 in range(2, max_row2 + 1):
        print(str(sheet1.cell(row1, 8).value) == str(sheet2.cell(row2, 1).value))
        if str(sheet1.cell(row1, 8).value) == str(sheet2.cell(row2, 1).value):
            sheet1.cell(row1, 18).value = sheet2.cell(row2, 2).value
            sheet1.cell(row1, 19).value = sheet2.cell(row2, 3).value

wb1.save('test12-28.xlsx')


