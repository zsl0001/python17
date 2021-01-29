import sys

sys.path.append("..")

from models import db, TMSOrderIndex, TMSDevice
from my_db.Models import Position, Log, Status
import time
import datetime
from mongoengine.queryset.visitor import Q
import openpyxl
from geopy.distance import geodesic

s_time = '2020-12-29 00:00:00'
e_time = '2021-01-04 00:00:00'


def get_location_time(t):
    if len(str(t)) > 19:
        t = str(t).split('.')[0]
    t = datetime.datetime.strptime(str(t), '%Y-%m-%d %H:%M:%S') + datetime.timedelta(hours=8)
    l_t = t.strftime("%Y-%m-%d %H:%M:%S")
    return t


def local2utc(local_st):
    """本地时间转UTC时间（-8: 00）"""
    if len(str(local_st)) > 19:
        local_st = str(local_st).split('.')[0]
    local_st = datetime.datetime.strptime(str(local_st), '%Y-%m-%d %H:%M:%S')
    time_struct = time.mktime(local_st.timetuple())
    utc_st = datetime.datetime.utcfromtimestamp(time_struct)
    return utc_st


def get_Utc_time(t):
    if len(str(t)) > 19:
        t = str(t).split('.')[0]
    a = datetime.datetime.strptime(str(t), "%Y-%m-%d %H:%M:%S")
    o = datetime.timedelta(hours=8)
    return a - o


def get_day3_time(t):
    if len(str(t)) > 19:
        t = str(t).split('.')[0]
    a = datetime.datetime.strptime(str(t), "%Y-%m-%d %H:%M:%S")
    a = a + datetime.timedelta(days=3)
    o = datetime.timedelta(hours=8)
    return a - o


def get_day4_time(t):
    if len(str(t)) > 19:
        t = str(t).split('.')[0]
    a = datetime.datetime.strptime(str(t), "%Y-%m-%d %H:%M:%S")
    a = a + datetime.timedelta(days=3)
    return a


def set_time(imei, start_time, sign_time=None):  # 将时间按3小时一次切割查询
    l = []
    dat = {'start_time': start_time,
           'end_time': sign_time,
           'imei': imei
           }
    t = str(start_time)
    while t <= str(sign_time):
        dat['start_time'] = t
        dat['end_time'] = str(datetime.datetime.strptime(t, '%Y-%m-%d %H:%M:%S') + datetime.timedelta(hours=3))
        t = str(datetime.datetime.strptime(t, '%Y-%m-%d %H:%M:%S') + datetime.timedelta(hours=3))
        l.append(dat.copy())
    l[-1]['end_time'] = sign_time
    return l


def get_position(data):
    my_list = []
    st = local2utc(str(data['start_time']))
    ed = local2utc(str(data['end_time']))
    res = Position.objects.filter((Q(time__gte=st) & Q(time__lte=ed) & Q(devId=str(data['imei'])))).order_by('time')
    for i in res:
        p_data = {
            'devId': i.devId,
            "longitude": i.longitude,
            "latitude": i.latitude,
            "time": get_location_time(i.time)
        }
        my_list.append(p_data)
    return my_list


def get_all_location(imei, start_time, end_time=None):
    sign_time = get_day4_time(end_time)
    l = set_time(imei=imei, start_time=start_time, sign_time=sign_time)
    m = []
    for i in l:
        mm = get_position(i)
        if len(mm) > 0:
            m = m + mm
    return m


# imei = '351608087073750'
s_time = '2020-12-30 00:00:00'
e_time = '2021-01-11 00:00:00'

# a = get_all_location(imei, s_time, e_time)
# print(a)
# my_ds = []
# for i in range(len(a)):
#     if i < len(a) - 1:
#         ds = (a[i + 1]['time'] - a[i]['time']).total_seconds()
#         my_ds.append(ds)
#         m_max = max(my_ds)
# # print(my_ds)
# # print(m_max)
# # print(my_ds.index(m_max))
# # print(a[my_ds.index(m_max)])
# a_s_time = a[my_ds.index(m_max)]['time'] - datetime.timedelta(hours=8)
# # print(a_s_time)
# imei = a[my_ds.index(m_max)]['devId']
# log_res = Log.objects.filter((Q(time__lte=a_s_time) & Q(imei=str(imei)))).order_by('time').limit(1)
# for k in log_res:
#     lec = str(k['content']).split(',')[-1]
wb = openpyxl.load_workbook('test1.xlsx')
sheet = wb['Sheet1']
max_row = sheet.max_row
for row in range(max_row):
    imei = sheet.cell(row + 2, 1).value
    a = get_all_location(imei, s_time, e_time)
    print(imei)
    my_ds = []
    for i in range(len(a)):
        if i < len(a) - 1:
            ds = (a[i + 1]['time'] - a[i]['time']).total_seconds()
            sheet.cell(row + 2, 5).value = a[i + 1]['time']
            sheet.cell(row + 2, 6).value = a[i]['time']
            my_ds.append(ds)
    if len(my_ds)==0:
        sheet.cell(row + 2, 2).value = '无历史数据'
        sheet.cell(row + 2, 3).value ='无历史数据'
    else:
        m_max = max(my_ds)
        # print(my_ds)
        # print(m_max)
        # print(my_ds.index(m_max))
        # print(a[my_ds.index(m_max)])
        a_s_time = a[my_ds.index(m_max)]['time'] - datetime.timedelta(hours=8)
        # print(a_s_time)
        imei = a[my_ds.index(m_max)]['devId']
        log_res = Log.objects.filter((Q(time__lte=a_s_time) & Q(imei=str(imei)))).order_by('time').limit(1)
        for k in log_res:
            lec = str(k['content']).split(',')[-1]
        sheet.cell(row + 2, 2).value = m_max
        sheet.cell(row + 2, 3).value = lec
        sheet.cell(row + 2, 4).value = str(a_s_time)
        sheet.cell(row + 2, 5).value = a[my_ds.index(m_max)]['time']
        sheet.cell(row + 2, 6).value = a[my_ds.index(m_max) -1]['time']
        sheet.cell(row + 2, 7).value = a[my_ds.index(m_max) + 1]['time']

wb.save('test2.xlsx')