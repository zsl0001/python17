import datetime
import json
from time import sleep

import requests
from openpyxl import load_workbook

# class DateEncoder(json.JSONEncoder):
#     def default(self, obj):
#         if isinstance(obj, datetime.datetime):
#             return obj.strftime('%Y-%m-%d %H:%M:%S')
#         elif isinstance(obj, date):
#             return obj.strftime("%Y-%m-%d")
#         else:
#             return json.JSONEncoder.default(self, obj)


url = 'http://106.14.17.157:7777/api/get_lg/'
url2 = 'http://106.14.17.157:7777/api/get_mp/'
wb = load_workbook('test0622.xlsx')
sheet = wb["Sheet1"]
l = []
data = {
    'code': 'qwer?1234!@#1312',
    'datalist': '',
}
for i in range(1, sheet.max_row + 1):
    Index_DeviceCode = sheet.cell(i, 2).value
    Index_PactCode = sheet.cell(i, 3).value
    l.append({'patcode': Index_PactCode, 'imei': Index_DeviceCode})
data['datalist'] = l
req = requests.post(url=url, data=json.dumps(data))
datalist = eval(req.text)

for i in range(1, sheet.max_row + 1):
    data2 = {
        'code': 'qwer?1234!@#1312',
        'start_time': '',
        'end_time': '',
        'imei': '',
    }
    Index_RealToTime = sheet.cell(i, 1).value
    Index_DeviceCode = sheet.cell(i, 2).value
    Index_PactCode = sheet.cell(i, 3).value
    Index_RealToTime2 = datetime.datetime.strptime(Index_RealToTime.split('.')[0], '%Y-%m-%d %H:%M:%S')
    Index_RealToTime2 = Index_RealToTime2 + datetime.timedelta(hours=-2)
    for m in datalist:
        to_lac = m[0]
        data2['start_time'] = str(Index_RealToTime2)
        data2['end_time'] = Index_RealToTime.split('.')[0]
        data2['imei'] = Index_DeviceCode
        print(data2)
        req2 = requests.post(url=url2, data=json.dumps(data2))
        da1 = eval(req2.text)
        c = 0
# 2020-06-19 14:36:05.993
# {'code': 'qwer?1234!@#1312', 'start_time': '2020-06-19 16:36:05', 'end_time': '2020-06-19 14:36:05', 'imei': '351608085026933'}
        for j in da1:
            t = str(j['latitude']) + ',' + str(j['longitude'])
            url = f"http://api.map.baidu.com/routematrix/v2/driving?output=json&origins={t}&destinations={to_lac}&ak=SMm8htpBXtu3Hd4n5XUsQwiUnMGvWdBU"
            res = requests.get(url)
            a = json.loads(res.text)
            print(a.get('result')[0]['distance'])
